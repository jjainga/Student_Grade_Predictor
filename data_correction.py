import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_wb(name,sheet_name,save):
    wb = xl.load_workbook(name)
    sheet = wb[sheet_name]
    fixing_price(sheet)

    wb.save(save)

#starting from 2 to ignore the headers
def fixing_price(sheet):
    for row in range(1,sheet.max_row +1):
        cell = sheet.cell(row,1).value
        row_list = cell.split(",")
        for i in range(len(row_list)):
            sheet.cell(row,i+1).value = row_list[i]


process_wb('student_mat.xlsx','student_mat','student_data.xlsx')